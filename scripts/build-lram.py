#!/usr/bin/env python3
# Preprocess the USDA Food Access Research Atlas 2019 (LRAM successor
# to FARA) into a small JSON keyed by 11-digit census tract GEOID.
#
# Why LRAM 2019 (not the 2015 or the newer SRAM):
#   - LRAM carries the LILATracts_Vehicle flag — Low-Income, Low-Access
#     using the vehicle-availability threshold, which is the signal that
#     matters for a food-access population that does not drive.
#   - SRAM uses different thresholds and drops the vehicle flag.
#
# Output: public/data/usda-lram-2019.json, filtered to NYC counties.
#
# Usage:
#   python3 scripts/build-lram.py [--xlsx path/to/FoodAccessResearchAtlasData2019.xlsx]
#
# If --xlsx is omitted, the script downloads it from USDA ERS.

import argparse
import json
import os
import sys
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Try: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(2)

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
OUT_PATH = REPO / "public" / "data" / "usda-lram-2019.json"
DEFAULT_XLSX_URL = "https://www.ers.usda.gov/sites/default/files/_laserfiche/DataFiles/80591/FoodAccessResearchAtlasData2019.xlsx"
CACHE_XLSX = REPO / ".venv-lram" / "FoodAccessResearchAtlasData2019.xlsx"

NYC_COUNTY_FIPS = {"005", "047", "061", "081", "085"}  # Bronx, Kings, NY, Queens, Richmond
NY_STATE_FIPS = "36"

# Columns we keep. Names come from the FARA/LRAM 2019 codebook; the sheet
# may capitalize differently across releases, so we lowercase for lookup.
WANTED = {
    "CensusTract": "geoid",
    "State": "state",
    "County": "county",
    "Urban": "urban",
    "Pop2010": "pop2010",
    "PovertyRate": "povertyRate",
    "MedianFamilyIncome": "medianFamilyIncome",
    "LowIncomeTracts": "lowIncome",
    "HUNVFlag": "hunvFlag",
    "LILATracts_Vehicle": "lilaVehicle",
    "LILATracts_1And10": "lila_1and10",
    "LILATracts_halfAnd10": "lila_halfand10",
    "TractLOWI": "tractLowIncomePop",
    "TractSNAP": "tractSnapHouseholds",
    "TractHUNV": "tractNoVehicleHouseholds",
    "TractSeniors": "tractSeniors",
    "TractKids": "tractKids",
}


def load_workbook_row_dicts(xlsx_path: Path):
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    # The main data sheet is usually the first, but the file also ships
    # variable-lookup and read-me sheets. Pick the one whose header row
    # contains "CensusTract".
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            continue
        header_norm = [str(h).strip() if h is not None else "" for h in header]
        lower = [h.lower() for h in header_norm]
        if "censustract" in lower:
            idx = {h: i for i, h in enumerate(header_norm)}
            return name, idx, rows
    raise SystemExit("Could not find a sheet with a CensusTract column")


def ensure_xlsx(user_path: str | None) -> Path:
    if user_path:
        p = Path(user_path).expanduser().resolve()
        if not p.exists():
            raise SystemExit(f"xlsx not found: {p}")
        return p
    if CACHE_XLSX.exists():
        print(f"  using cached xlsx: {CACHE_XLSX}")
        return CACHE_XLSX
    print(f"  downloading {DEFAULT_XLSX_URL}")
    CACHE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(DEFAULT_XLSX_URL, headers={"User-Agent": "nypl-hackathon/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp, open(CACHE_XLSX, "wb") as f:
        f.write(resp.read())
    print(f"  saved {CACHE_XLSX} ({CACHE_XLSX.stat().st_size} bytes)")
    return CACHE_XLSX


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="Path to FoodAccessResearchAtlasData2019.xlsx (else download)")
    args = ap.parse_args()

    xlsx_path = ensure_xlsx(args.xlsx)
    print(f"Reading {xlsx_path.name}")
    sheet_name, header_idx, rows = load_workbook_row_dicts(xlsx_path)
    print(f"  sheet: {sheet_name}")

    missing = [c for c in WANTED if c not in header_idx]
    if missing:
        print(f"  warn: columns not found in sheet: {missing}")

    out = {}
    total = 0
    nyc = 0
    lila_vehicle_count = 0
    for row in rows:
        total += 1
        if row is None:
            continue
        raw_geoid = row[header_idx["CensusTract"]] if "CensusTract" in header_idx else None
        if raw_geoid is None:
            continue
        # GEOID is numeric in the xlsx; pad to 11 chars.
        geoid = str(int(raw_geoid)).zfill(11) if isinstance(raw_geoid, (int, float)) else str(raw_geoid).strip()
        state = geoid[:2]
        county = geoid[2:5]
        if state != NY_STATE_FIPS or county not in NYC_COUNTY_FIPS:
            continue
        nyc += 1
        rec = {"geoid": geoid}
        for src, dst in WANTED.items():
            if src == "CensusTract":
                continue
            if src not in header_idx:
                continue
            v = row[header_idx[src]]
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            rec[dst] = v
        if rec.get("lilaVehicle") in (1, "1", True):
            lila_vehicle_count += 1
        out[geoid] = rec

    payload = {
        "_meta": {
            "dataset": "USDA Food Access Research Atlas, 2019 data (LRAM)",
            "publisher": "USDA Economic Research Service",
            "source_file": xlsx_path.name,
            "source_url": DEFAULT_XLSX_URL,
            "vintage_label": "USDA LRAM, 2019 data",
            "filter": f"state={NY_STATE_FIPS}, counties in {sorted(NYC_COUNTY_FIPS)} (NYC 5 boroughs)",
            "flag_note": "lilaVehicle=1 means Low-Income Low-Access using the vehicle-availability threshold — the flag that matters for a population that does not drive.",
            "tract_count": len(out),
        },
        "tracts": out,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(payload, f, separators=(",", ":"))
        f.write("\n")

    print(f"\n  rows scanned:            {total}")
    print(f"  NYC tracts kept:         {nyc}")
    print(f"  LILA-vehicle tracts:     {lila_vehicle_count}")
    print(f"  wrote {OUT_PATH} ({OUT_PATH.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
